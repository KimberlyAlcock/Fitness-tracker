from flask import Flask, render_template, request, jsonify
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('planks.html')

# Pulls data from the excel file practice-tracker.xlsx
print("app starting")
@app.route('/get_excel_data')
def get_excel_data():
    # Load the Excel file
    wb = openpyxl.load_workbook('./data/practice-tracker.xlsx')
    sheet = wb.active

    # Read data from Excel file
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Return data as JSON
    print(data)
    print(jsonify(data))
    return jsonify(data)
    
print("app finishing")
if __name__ == '__main__':
    app.run(debug=True)

# Adds data to the excel file practice-tracker.xlsx
@app.route('/submit', methods=['POST'])
def submit():
    data = request.form['data']

    # Open the Excel file
    wb = openpyxl.load_workbook('./data/practice-tracker.xlsx')
    ws = wb.active

    # Append data to the next empty row
    row = (data,)
    ws.append(row)

    # Save the changes
    wb.save('practice-tracker.xlsx')

    return 'Data submitted successfully!'

if __name__ == '__main__':
    app.run(debug=True)